from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
ws = wb.active
ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New Title"]
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

c = ws['A4']

#wb.save('new_sample.xlsx')

import openpyxl
openpyxl.worksheet.worksheet.Worksheet.